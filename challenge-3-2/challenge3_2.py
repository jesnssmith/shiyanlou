# -*- coding: utf-8 -8-

from openpyxl import load_workbook
from openpyxl import Workbook
import datetime

def combine():
	wb = load_workbook('courses.xlsx')
	ws1 = wb[wb.sheetnames[0]]
	ws2 = wb[wb.sheetnames[1]]
	ws3 = wb.create_sheet(title='combine')
	titles = [x.value for x in ws1['B']]
	for row in range(1,ws1.max_row+1):
		for col in range(1,ws1.max_column+1):
			
			ws3.cell(column = col, row = row).value = ws1.cell(column = col, row = row).value

	for i in ws2.rows:
		ws3.cell(column = 4, row = titles.index(i[1].value)+1).value = i[2].value

	for row in ws3.rows:
		for cell in row:
			print(cell.value,end = '  ')
		print()

	# wb.save('courses.xlsx')

def split():
	wb = load_workbook('courses.xlsx')
	ws = wb['combine']
	years={}
	i=0
	for row in ws.rows:
		i+=1
		if i == 1:
			row_head=[x.value for x in row]
			continue
		row_mirror=[x.value for x in row]
		if years.get(row[1].year):
			years.get(row[1].year).active.append(row_mirror)
		else:
			years[row[1].year] = Workbook().active.append(row_head)
			years[row[1].year] = Workbook().active.append(row_mirror)

if __name__ == '__main__':
	combine()